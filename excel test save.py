import cv2
import numpy as np
import time
import os
import math
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def calculate_contour_metrics(contours):
    if not contours:
        print("No contours found")
        return None
    
    #contours_np = [np.array(contour) for contour in contours]
    
    cnt = max(contours, key=cv2.contourArea)
    
    area_original = cv2.contourArea(cnt)
    perimeter_original = cv2.arcLength(cnt, True)
    
    if area_original <= 1e-6 or perimeter_original <= 1e-6:
        print(f"Invalid contour measurements: area={area_original}, perimeter={perimeter_original}")
        return None
    
    circularity_original = float(2 * math.sqrt((math.pi) * area_original)) / perimeter_original
    
    hull = cv2.convexHull(cnt)
    
    area_hull = cv2.contourArea(hull)
    perimeter_hull = cv2.arcLength(hull, True)
    
    if area_hull <= 1e-6 or perimeter_hull <= 1e-6:
        print(f"Invalid hull measurements: area={area_hull}, perimeter={perimeter_hull}")
        return None
    
    circularity_hull = float(2 * math.sqrt((math.pi) * area_hull)) / perimeter_hull
    
    area_ratio = area_hull / area_original
    circularity_ratio = circularity_hull / circularity_original

    return {
        "area_original": area_original,
        "area_hull": area_hull,
        "area_ratio": area_ratio,
        "circularity_original": circularity_original,
        "circularity_hull": circularity_hull,
        "circularity_ratio": circularity_ratio,
        "contour": cnt,
        "hull": hull
    }

def process_image(image_path, blurred_bg):
    print(f"Processing image: {image_path}")
    start_time = time.perf_counter()
    
    image = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
    
    kernel = cv2.getStructuringElement(cv2.MORPH_CROSS, (3, 3))
    blurred = cv2.GaussianBlur(image, (5, 5), 0)

    bg_sub = cv2.subtract(blurred_bg, blurred)
    _, binary = cv2.threshold(bg_sub, 10, 255, cv2.THRESH_BINARY)

    dilate1 = cv2.dilate(binary, kernel, iterations=1)
    erode1 = cv2.erode(dilate1, kernel, iterations=1)
    #erode2 = cv2.erode(erode1, kernel, iterations=1)
    #dilate2 = cv2.dilate(erode2, kernel, iterations=2)

    edges = cv2.Canny(erode1, 50, 150)
    contours, _ = cv2.findContours(edges, cv2.RETR_LIST, cv2.CHAIN_APPROX_NONE)
    results = calculate_contour_metrics(contours)
    
    end_time = time.perf_counter()
    dif_time = end_time - start_time
        # Prepare an image to draw the contours
    contour_image = np.zeros_like(image)
    # Show the resulting image
    cv2.drawContours(contour_image, contours, -1, 255, 1)
    cv2.imshow('Processed Image', contour_image)
    cv2.waitKey(1)
    #cv2.destroyAllWindows()
    print(f"Finished processing {image_path} in {dif_time:.2f} seconds")
    
    return contours, results, image, dif_time

def process_all_images(folder_path, background_path):
    results = []
    filename = sorted([f for f in os.listdir(folder_path) if f.endswith('.tiff')])
    #filenames_with_prefix = [f"{i:04d}_{f}" for i, f in enumerate(sorted(filename), 1)]
    for filename in filename:
        image_path = os.path.join(folder_path, filename)
        try:
            contours, result, _, dif_time = process_image(image_path, background_path)
            if result:
                result['filename'] = filename  # 使用原始文件名
                result['processing_time'] = dif_time
                results.append(result)
            else:
                print(f"No valid results for {filename}")
        except Exception as e:
            print(f"Error processing {filename}: {str(e)}")
    
    return results


def save_to_excel(results, output_file):
    print("Saving results to Excel...")
    
    # 檢查文件是否存在
    if os.path.exists(output_file):
        # 如果文件存在，加載現有的工作簿
        wb = load_workbook(output_file)
        ws = wb.active
    else:
        # 如果文件不存在，創建新的工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"

        # 寫入表頭
        headers = ["Filename", "Test", "Area Original", "Area Hull", "Area Ratio", 
                   "Circularity Original", "Circularity Hull", "Circularity Ratio",
                   "Processing Time"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

    # 寫入數據
    for result in results:
        filename = result['filename']
        
        # 查找文件名或創建新的行
        file_row = None
        for row in range(2, ws.max_row + 1, 10):
            if ws.cell(row=row, column=1).value == filename:
                file_row = row
                break
        
        if file_row is None:
            # 新文件名，創建新的10行區塊
            file_row = ws.max_row + 1
            ws.merge_cells(start_row=file_row, start_column=1, end_row=file_row+9, end_column=1)
            cell = ws.cell(row=file_row, column=1, value=filename)
            cell.alignment = Alignment(vertical='top')

        # 找到下一個可用的 test 行
        for i in range(10):
            test_cell = ws.cell(row=file_row+i, column=2)
            if test_cell.value is None or test_cell.value == "":
                test_cell.value = f"dilate erode各僅一次"
                result_row = file_row + i
                break
        else:
            print(f"Warning: No more test slots available for {filename}")
            continue

        # 寫入結果
        ws.cell(row=result_row, column=3, value=result['area_original'])
        ws.cell(row=result_row, column=4, value=result['area_hull'])
        ws.cell(row=result_row, column=5, value=result['area_ratio'])
        ws.cell(row=result_row, column=6, value=result['circularity_original'])
        ws.cell(row=result_row, column=7, value=result['circularity_hull'])
        ws.cell(row=result_row, column=8, value=result['circularity_ratio'])
        ws.cell(row=result_row, column=9, value=result['processing_time'])


    # 調整列寬
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 20
    if result and all(key in result for key in ['area_original', 'area_hull', 'area_ratio', 'circularity_original', 'circularity_hull', 'circularity_ratio', 'processing_time']):
        results.append(result)
    else:
        print(f"Incomplete or invalid result for {filename}")
    wb.save(output_file)
    print(f"Excel file saved: {output_file}")


def main():
    folder_path = 'Test_images/Slight under focus'
    background_path = 'Test_images/Slight under focus/background.tiff'
    output_file = 'C:/Users/USER/RT-DC-master/contour_metrics_results.xlsx'
    background = cv2.imread(background_path, cv2.IMREAD_GRAYSCALE)
    if background is None:
        print(f"Failed to read image or background: {background_path}")
        return None, None, None, 0
    blurred_bg = cv2.GaussianBlur(background, (5, 5), 0)
    
    start_time = time.time()
    results = process_all_images(folder_path, blurred_bg)
    save_to_excel(results, output_file)
    end_time = time.time()
    
    print("Excel file saved at:", os.path.abspath(output_file))
    print(f"Total processing time: {end_time - start_time:.2f} seconds")
    print(f"Results saved to {output_file}")

if __name__ == "__main__":
    main()